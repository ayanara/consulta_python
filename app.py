import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep


# 1 - Entrar na planilha e extrair o cpf 
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_cliente = planilha_clientes['Sheet1']
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_cliente.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf , vencimento = linha
    # 2 - Entrar no site e usar o cpf para pesquisar o status do cliente\
    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    # 3 - Verificar se está em dia ou atrasado
    botao_pesquisa = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisa.click()
    sleep(4)
    
    campo_status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    if campo_status.text == 'em dia':
        # 4 - Verificar se esta em dia, pegar a data do pagamento
        data_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia'])
    else:
        # 5 - Caso contrario, colocar status pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_clientes['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente'])
